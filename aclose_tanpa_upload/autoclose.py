import zipfile
from selenium import webdriver
from undetected_chromedriver import Chrome
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions as EC
import sys
import datetime
import re
from time import *
import os
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.proxy import Proxy, ProxyType
import random
import secrets
import string
import requests
import json
import pandas as pd
import psutil
import openpyxl
import json,urllib.request
import pickle
from PIL import Image


#DATE TIME GMT DAN LOCAL
# print("\nGMT: "+time.strftime("%a, %d %b %Y %I:%M:%S %p %Z", time.gmtime()))
# print("Local: "+strftime("%a, %d %b %Y %I:%M:%S %p %Z\n"))
stop = False
def shot_api(incident_id, atm_id, problem):
	print('shot_api ',incident_id,atm_id,problem)
	url = 'https://boss.citius.co.id/api/create-ticket'
	myobj = {'incident_id':incident_id, 'atm_id':atm_id, 'problem':problem}
	x = requests.post(url, data = myobj)
	x.close()
	return x.text

def element_presence(by,by_val,time, driver):
	element_present = EC.presence_of_element_located((by, by_val))
	try:
		WebDriverWait(driver, time).until(element_present)
	except Exception as e:
		print (e)
	else:
		pass

def run():

	dataframe = openpyxl.load_workbook('doc.xlsx')
	dataframe1 = dataframe.active
	conti_last_proc = False
	last_proc = ""
	try:
		last_proc = open("last_line_proccessed.txt","r").readline()
		conti_last_proc = True
	except Exception as e:
		pass

	# for row in range(0, dataframe1.max_row):
	# 	print(dataframe1[row])
	# 	for col in dataframe1.iter_cols(1, dataframe1.max_column):
	# 		print(col[row].value)
	# 	input("enter")
	# tab = sys.argv[1]
	chrome_options = Options()
	chrome_options.add_argument("--incognito")
	chrome_options.add_argument('--headless')
	#chrome_options.add_argument('--no-sandbox')
	#chrome_options.add_argument('--proxy-server='+input_proxy.split("-")[0])
	driver_d = Chrome(options=chrome_options)
	driver_d.maximize_window()
	has_cookie = 0
	driver_d.get("https://app.slmugmandiri.co.id/sistrack_new/")
	login(driver_d)
	line_no = 0
	for data in dataframe1:
		try:
			line_no += 1
			ebs = data[2].value
			reason = data[3].value
			if conti_last_proc:
				if str(line_no) == last_proc:
					conti_last_proc = False
				else:
					print("SKIPPING LINE",line_no)
					continue
			if "-" not in ebs:
				continue
			print("PROCCESSING ",ebs)
			driver_d.get("https://app.slmugmandiri.co.id/sistrack_new/Home")
	
			element_presence(By.XPATH, "/html/body/div[2]/div/div/div[2]/div/div[4]/div/div/div/div/div[1]/div[2]/div/label/input", 30, driver_d)
			driver_d.find_element(By.XPATH, "/html/body/div[2]/div/div/div[2]/div/div[4]/div/div/div/div/div[1]/div[2]/div/label/input").send_keys(ebs+"\n")

			xpath_status_text = "/html/body/div[2]/div/div/div[2]/div/div[4]/div/div/div/div/div[2]/div/table/tbody/tr[1]/td[3]/h3"
			while True:
				sleep(1)
				try:
					driver_d.find_element(By.XPATH, "/html/body/div[2]/div/div/div[2]/div/div[4]/div/div/div/div/div[2]/div/table/tbody/tr[2]/td[1]")
					driver_d.find_element(By.XPATH, xpath_status_text)
					print("ADA")
				except Exception as e:
					break

			element_presence(By.XPATH, xpath_status_text, 30, driver_d)

			if driver_d.find_element(By.XPATH, xpath_status_text).text != "CLOSED":
				driver_d.find_element(By.XPATH, "/html/body/div[2]/div/div/div[2]/div/div[4]/div/div/div/div/div[2]/div/table/tbody/tr/td[1]/a").click()
				element_presence(By.XPATH, "/html/body/div[2]/div/div/div[2]/div[4]/div[2]/form/div/div[1]/table/tbody/tr[1]/td[2]", 30, driver_d)
				sleep(2)
				try:
					btn_suspen = driver_d.find_element(By.XPATH, "/html/body/div[2]/div/div/div[2]/div[3]/div/div/a");
					if "Open Suspend" in btn_suspen.text:
						driver_d.find_element(By.XPATH, "/html/body/div[2]/div/div/div[2]/div[3]/div/div/a").click()
						element_presence(By.XPATH, "/html/body/div[2]/div/div/div[2]/div[1]/div/div", 30, driver_d)
				except Exception as e:
					pass

				driver_d.find_element(By.XPATH, "/html/body/div[2]/div/div/div[2]/div[4]/div[2]/form/div/div[1]/table/tbody/tr[1]/td[2]/input").send_keys("-")

				driver_d.find_element(By.XPATH, "/html/body/div[2]/div/div/div[2]/div[4]/div[2]/form/div/div[1]/table/tbody/tr[2]/td[2]/select").click()
				driver_d.find_element(By.XPATH, "/html/body/div[2]/div/div/div[2]/div[4]/div[2]/form/div/div[1]/table/tbody/tr[2]/td[2]/select/option[2]").click()

				driver_d.find_element(By.XPATH, "/html/body/div[2]/div/div/div[2]/div[4]/div[2]/form/div/div[1]/table/tbody/tr[3]/td[2]/input").send_keys("-")

				driver_d.find_element(By.XPATH, "/html/body/div[2]/div/div/div[2]/div[4]/div[2]/form/div/div[1]/table/tbody/tr[4]/td[2]/select").click()
				driver_d.find_element(By.XPATH, "/html/body/div[2]/div/div/div[2]/div[4]/div[2]/form/div/div[1]/table/tbody/tr[4]/td[2]/select/option[2]").click()

				driver_d.find_element(By.XPATH, "/html/body/div[2]/div/div/div[2]/div[4]/div[2]/form/div/div[1]/table/tbody/tr[5]/td[2]/input").send_keys("-")

				driver_d.find_element(By.XPATH, "/html/body/div[2]/div/div/div[2]/div[4]/div[2]/form/div/div[1]/table/tbody/tr[6]/td[2]/select").click()
				driver_d.find_element(By.XPATH, "/html/body/div[2]/div/div/div[2]/div[4]/div[2]/form/div/div[1]/table/tbody/tr[6]/td[2]/select/option[2]").click()

				driver_d.find_element(By.XPATH, "/html/body/div[2]/div/div/div[2]/div[4]/div[2]/form/div/div[1]/table/tbody/tr[7]/td[2]/input").send_keys("-")

				driver_d.find_element(By.XPATH, "/html/body/div[2]/div/div/div[2]/div[4]/div[2]/form/div/div[1]/table/tbody/tr[8]/td[2]/select").click()
				driver_d.find_element(By.XPATH, "/html/body/div[2]/div/div/div[2]/div[4]/div[2]/form/div/div[1]/table/tbody/tr[8]/td[2]/select/option[2]").click()
				driver_d.find_element(By.XPATH, "/html/body/div[2]/div/div/div[2]/div[4]/div[2]/form/div/div[2]/table/tbody/tr[6]/td/input").send_keys("-")

				driver_d.find_element(By.XPATH, "/html/body/div[2]/div/div/div[2]/div[4]/div[2]/form/div/div[2]/table/tbody/tr[8]/td/textarea").send_keys(reason)

				driver_d.find_element(By.XPATH, "/html/body/div[2]/div/div/div[2]/div[4]/div[2]/form/div/div[3]/button").click()
				element_presence(By.XPATH, "/html/body/div[2]/div/div/div[2]/div[1]/div/div", 30, driver_d)
				sleep(2)
				driver_d.refresh()
				element_presence(By.XPATH, "/html/body/div[2]/div/div/div[2]/div[4]/div[2]/form/div/div[1]/table/tbody/tr[1]/td[2]", 30, driver_d)

				driver_d.find_element(By.XPATH, "/html/body/div[2]/div/div/div[2]/div[4]/div[1]/div[2]/table/tbody/tr[4]/td[2]/div/div[1]/select").click()
				driver_d.find_element(By.XPATH, "/html/body/div[2]/div/div/div[2]/div[4]/div[1]/div[2]/table/tbody/tr[4]/td[2]/div/div[1]/select/option[6]").click()
				driver_d.find_element(By.XPATH, "/html/body/div[2]/div/div/div[2]/div[4]/div[1]/div[2]/table/tbody/tr[4]/td[2]/div/div[2]/button").click()
				element_presence(By.XPATH, "/html/body/div[2]/div/div/div[2]/div[1]/div/div", 30, driver_d)
			open("last_line_proccessed.txt","w").writelines(str(line_no))

			if int(line_no) == int(dataframe1.max_row):
				stop = True
				print("sini", stop)
				driver_d.quit()
				print("DONE")
				open("last_line_proccessed.txt","w").writelines("1")
				open("log_status.txt","w").writelines("DONE")
				break

		except Exception as e:
			open("ticket_error.txt","a").writelines(ebs+"-EXCEPTION\n"+str(e)+"\n\n")
			driver_d.quit()
			break

			
# def check_whitelist_sender_email(email):
# 	url = 'https://boss.citius.co.id/api/check-whitelist-email'
# 	myobj = {'email':email}
# 	while True:
# 		try:
# 			x = requests.post(url, data = myobj)
# 			x.close()
# 			if 'ok' in x.text:
# 				return True;
# 			return False;
# 		except Exception as e:
# 			print('EERORRR', e)

def check_whitelist_sender_email(email):
	whitelist_email = ['ccugmandiri@gmail.com', 'citiusdispatcher@gmail.com']
	for x in whitelist_email:
		if email == x:
			return True
	return False

def download(url, name):
	while True:
		try:
			response = requests.get(url, verify=False, timeout=10)
			print("OKEEEE")
			open(name, "wb").write(response.content)
			break
		except Exception as e:
			print("ULANGGGG")
			sleep(5)

def shot_url(url):
	while True:
		try:
			myobj = {}
			x = requests.get(url, verify=False, timeout=10)
			x.close()
			break
		except Exception as e:
			print("ULANGGGG")
			sleep(5)

def get_data_api(url):
	while True:
		try:
			x = requests.get(url, verify=False, timeout=10)
			x.close()
			return json.loads(x.text)
			break
		except Exception as e:
			print("ULANGGGG")
			sleep(5)

def post_data_api(url):
	while True:
		try:
			x = requests.post(url, data = myobj, verify=False, timeout=10)
			x.close()
			return json.loads(x.text)
			break
		except Exception as e:
			print("ULANGGGG")
			sleep(5)

def login(driver_d):
	element_presence(By.XPATH, "/html/body/form/div/input[1]", 30, driver_d)
	sleep(3)
	try:
		driver_d.find_element(By.XPATH, "/html/body/form/div/input[1]").send_keys("CC CTS 3")
		element_presence(By.XPATH, "/html/body/form/div/input[2]", 30, driver_d)
		sleep(2)
		driver_d.find_element(By.XPATH, "/html/body/form/div/input[2]").send_keys("123456\n")
	except Exception as e:
		print(e)


while True:
	open("log_status.txt","w").writelines("ONPROGRESS")
	try:
		run()
		progres = open("log_status.txt","r").readline()
		if progres == "DONE":
			break
	except Exception as e:
		print(e)
		input("error")